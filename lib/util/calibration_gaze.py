import sys
sys.path.insert(0, 'lib/')
import os
import torch
import numpy as np
import torch.utils.data as Data
import torch.nn as nn
from tqdm import trange, tqdm
from data_loader.data_loader import Personal_Train_Data, Personal_Test_Data
from models.gaze_estimator_model import DenseNet
import yaml
import openpyxl

with open('config.yaml', 'r') as yml:
	config = yaml.safe_load(yml)
    

def loop_iter(dataloader):
    while True:
        for data in iter(dataloader):
            yield data

def get_train_data_loaders(batch_size, data_name):
    num = 3
    train_data = Personal_Train_Data(data_name)
    data_loader = Data.DataLoader(train_data, batch_size=batch_size, shuffle=False, pin_memory=True, drop_last=False, num_workers=num)
    data_iter = loop_iter(data_loader)

    return data_iter

def get_test_data_loaders(batch_size, data_name):
    num = 3
    test_data = Personal_Test_Data(data_name)
    data_loader = Data.DataLoader(test_data, batch_size=batch_size, shuffle=False, pin_memory=True, drop_last=False, num_workers=num)
    data_iter = loop_iter(data_loader)

    return data_iter

def gaze_diff_func(gt_gaze, gaze, bias):
    gt_gaze = gt_gaze.cuda()
    gaze = gaze.cuda()
    cos = nn.CosineSimilarity()
    '''Normalize gaze direction'''
    gaze = torch.squeeze(gaze)
    gaze = gaze + bias
    difference = torch.mean(gt_gaze - gaze, dim=0)
    cos_sim = cos(gt_gaze, gaze)
    cos_sim = cos_sim.to('cpu').detach().numpy().copy()
    cos_sim = cos_sim.astype(np.float32)
    cos_sim_rad = np.arccos(cos_sim)

    diff_gaze = cos_sim_rad * (180 / np.pi)

    return diff_gaze, difference

def gaze_estimate(data_name, Densenet, batch_size, bias):
    print(data_name)
    img_dir =  data_name
    img_max = len(os.listdir(img_dir))
    difference_list = []
    diff_gaze_array = np.empty(0)
    data_loader = get_test_data_loaders(batch_size, img_dir)

    for i in tqdm(range(int(img_max / batch_size))):
        test_img, gt_gaze, _ = next(data_loader)
        img_tensor = test_img.unsqueeze(0).cuda()
        img_tensor = torch.squeeze(img_tensor, dim=0)
        pred_gaze = Densenet(img_tensor)
        pred_gaze_normalized = torch.nn.functional.normalize(pred_gaze, dim=2)

        gt_gaze = np.array(gt_gaze)
        gt_gaze = gt_gaze.T
        gt_gaze = gt_gaze[:,1:4]
        gt_gaze = torch.from_numpy(gt_gaze.astype(np.float32)).clone()
        bias = bias.cuda()
        # import pdb;pdb.set_trace()
        gaze_difference, difference = gaze_diff_func(gt_gaze, pred_gaze_normalized, bias)

        diff_gaze_array = np.append(diff_gaze_array, gaze_difference)
        difference_list.append(difference)
    
    mean_diff_gaze = np.mean(diff_gaze_array)
    diff_tensor = torch.reshape(torch.cat(difference_list, dim=0),(-1,3))
    bias_output = torch.mean(diff_tensor, dim=0).cpu()
    # print('test_data_bias:',bias_output)
    # print('test_mean_diff_gaze:', mean_diff_gaze)

    return mean_diff_gaze, bias, bias_output


def measure_error_average(train_data_name, Densenet, batch_size):

    Densenet.eval()
    for param in Densenet.parameters():
        param.requires_grad = False


    img_dir =  '../data/' + train_data_name + '/'
    print(img_dir)
    img_max = len(os.listdir(img_dir))
    # import pdb;pdb.set_trace()
    test_data_name = img_dir[:-5] + '_test' + img_dir[-5:]
    bias_zero = torch.tensor([0.00, 0.00, 0.00]).cuda()
    bias = None
    difference_list = []
    diff_gaze_array = np.empty(0)
    data_loader = get_train_data_loaders(batch_size, train_data_name)

    for i in tqdm(range(int(img_max / batch_size))):
        test_img, gt_gaze, _ = next(data_loader)
        img_tensor = test_img.unsqueeze(0).cuda()
        img_tensor = torch.squeeze(img_tensor, dim=0)
        # import pdb;pdb.set_trace()
        pred_gaze = Densenet(img_tensor)

        pred_gaze_normalized = torch.nn.functional.normalize(pred_gaze, dim=2)

        gt_gaze = np.array(gt_gaze)
        gt_gaze = gt_gaze.T
        gt_gaze = gt_gaze[:,1:4]
        gt_gaze = torch.from_numpy(gt_gaze.astype(np.float32)).clone()

        # import pdb;pdb.set_trace()
        # print(gt_gaze.dtype, pred_gaze.dtype, bias_zero.dtype)
        gaze_difference, difference = gaze_diff_func(gt_gaze, pred_gaze_normalized, bias_zero)
        # diff_gaze_array = np.append(diff_gaze_array, diff_gaze)
        # mean_diff_gaze = np.mean(diff_gaze_array)
        # print('Validation Mean Error: ' + str(mean_diff_gaze))

        difference_list.append(difference)
        diff_gaze_array = np.append(diff_gaze_array, gaze_difference)

    mean_diff_gaze = np.mean(diff_gaze_array)
    diff_tensor = torch.reshape(torch.cat(difference_list, dim=0),(-1,3))
    bias = torch.mean(diff_tensor, dim=0).cpu()
    print('train_data_bias:',bias)
    # print('train_mean_diff_gaze:', mean_diff_gaze)

    gaze_diff, bias, bias_calib = gaze_estimate(test_data_name, Densenet, batch_size, bias)
        
    return gaze_diff, bias, bias_calib, mean_diff_gaze

# def build_test_network(cnt):
#     in_channels = 1
#     out_channels = 64
#     stride = 2	
#     checkpoints = os.listdir(config['checkpoint_path']['pathname'])
#     # self.Reg = Regressor(3, 32, 34).cuda()
#     Densenet = DenseNet(num_init_features=out_channels).cuda()

#     if torch.cuda.device_count() > 1:
#         print("Let's use", torch.cuda.device_count(), "GPUs!")
#         device_id = config['gpu_num']['num']
#         Densenet = torch.nn.DataParallel(Densenet, device_ids=device_id)
#         Densenet = Densenet.cuda()
    
#     densenet_checkpoints = [ckpt for ckpt in checkpoints if 'Densenet_' == ckpt[:9]]
#     densenet_checkpoints.sort(key=lambda x: int(x[9:-4]), reverse=False)

#     if len(os.listdir(config['checkpoint_path']['pathname'])) == 0:
#         print("No Previous Weights Found. Building and Initializing new Model")
#         self.current_step = 0
#         return False

#     Densenet.load_state_dict(torch.load(os.path.join(config['checkpoint_path']['pathname'], densenet_checkpoints[cnt])), strict=False)
#     print(os.path.join(config['checkpoint_path']['pathname'], densenet_checkpoints[cnt]))

#     return Densenet


def calibrated_gaze(Densenet, steps):
    batch_size = 250
    row_cnt = int(steps/config['log_interval']['num'])
    # row_cnt = int(steps/1)
    # import pdb;pdb.set_trace()
    test_data_list = []
    # bias_list = []
    gaze_diff_array = np.empty(0)
    gaze_filename =  "../result/test/mpiigaze/mpiigaze_calib_test.txt" 

    for i in range(15):
        subject_name = 'mpiigaze_personal/p' + str(i).zfill(2)
        test_data_list.append(subject_name)
        # print(subject_name)

    for test_data_name in test_data_list:
        print('test_data_name:',test_data_name)
        gaze_diff, bias, bias_calib, non_bias_gaze_diff = measure_error_average(test_data_name, Densenet, batch_size)
        gaze_diff_array = np.append(gaze_diff_array, gaze_diff)
        # bias_list.append(bias) 
        bias_txt_name = '../result/test/mpiigaze/' + test_data_name[-3:] + '_bias.xlsx'


        if not os.path.exists(bias_txt_name):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "test"
            wb.save(bias_txt_name)

        wb = openpyxl.load_workbook(bias_txt_name)
        ws = wb.active
        sheet = wb['test']  
        sheet.cell(column=1,row=row_cnt).value = float(bias[0])
        sheet.cell(column=2,row=row_cnt).value = float(bias[1])
        sheet.cell(column=3,row=row_cnt).value = float(bias[2])
        sheet.cell(column=5,row=row_cnt).value = float(bias_calib[0])
        sheet.cell(column=6,row=row_cnt).value = float(bias_calib[1])
        sheet.cell(column=7,row=row_cnt).value = float(bias_calib[2])
        sheet.cell(column=9,row=row_cnt).value = float(non_bias_gaze_diff)
        sheet.cell(column=10,row=row_cnt).value = float(gaze_diff)
        wb.save(bias_txt_name)

        # f = open(bias_txt_name, 'a')
        # f.write(str(bias)+ '\n')
        # f.close()


        # import pdb;pdb.set_trace()
    
        mean_diff_gaze = np.mean(gaze_diff_array,0)
        # result.append(mean_diff_gaze)

        '''Output Gaze Difference .txt'''
        # subject_bias_gaze = '../result/test/mpiigaze/' + test_data_name[-3:] + '_bias_gaze.text'
        # subject_non_bias_gaze = '../result/test/mpiigaze/' + test_data_name[-3:] + '_non_bias_gaze.text'

        # if not os.path.exists(subject_bias_gaze):
        #     f_bias = open(subject_bias_gaze,'w')
        #     f_bias.close()
        # if not os.path.exists(subject_non_bias_gaze):
        #     f_non_bias = open(subject_non_bias_gaze,'w')
        #     f_non_bias.close()
            
        # f_bias = open(subject_bias_gaze, 'a')
        # f_bias.write(str(gaze_diff) + '\n')
        # f_bias.close()

        # f_non_bias = open(subject_non_bias_gaze, 'a')
        # f_non_bias.write(str(non_bias_gaze_diff) + '\n')
        # f_non_bias.close()
		
    f = open(gaze_filename, 'a')
    f.write(str(mean_diff_gaze) + '\n')
    f.close()
    
    return mean_diff_gaze
    



# if __name__ == '__main__':
#     batch_size = 250
#     checkpoints = os.listdir(config['checkpoint_path']['pathname'])
#     checkpoints_num = len([ckpt for ckpt in checkpoints if 'Densenet_' == ckpt[:9]])
#     Densenet = None
#     result = []

#     for cnt in range(checkpoints_num):
#         gaze_diff_array = np.empty(0)
#         gaze_diff_list = []
#         bias_list = []
#         test_data_list = []
#         Densenet = build_test_network(cnt)

#         for i in range(15):
#             subject_name = 'mpiigaze_personal/p' + str(i).zfill(2)
#             test_data_list.append(subject_name)
#             # print(subject_name)

#         for test_data_name in test_data_list:
#             print('test_data_name:',test_data_name)
#             gaze_diff, bias = measure_error_average(test_data_name, Densenet, batch_size)
#             gaze_diff_array = np.append(gaze_diff_array, gaze_diff)
#             bias_list.append(bias) 

#         # import pdb;pdb.set_trace()
#         mean_diff_gaze = np.mean(gaze_diff_array,0)
#         result.append(mean_diff_gaze)

#     np.savetxt('../result/test/mpiigaze/mpiigaze_bias_diff.txt', result) 
            # np.savetxt('../result/test/mpiigaze/mpiigaze_bias')

